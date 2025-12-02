#!/usr/bin/env python

"""
============================================================
Simpy Simulation UI – Main Screen and Settings Interface
============================================================

This script provides a graphical user interface (GUI) for configuring
and launching a Simpy-based simulation. It uses Tkinter to create a
main window with interactive sliders for setting various simulation
parameters.

Main Features:
--------------
- Main window with buttons to:
  - Open a settings window for simulation parameters
  - Launch a separate station creator script (UI_1.py)
- Dynamic creation of sliders for parameter input
- Saving user-defined parameters to a JSON file for later use
- Scrollable interface for better usability with many parameters

Files:
------
- simulation_data.json: Stores the user-defined simulation parameters

Note:
-----
Ensure that the file "UI_1.py" is located in the same directory as
this script so that the station creator can be launched correctly.

Author: Dr. Patric Schubert
Date: July 2025
"""


import tkinter as tk
from tkinter import ttk, font
from ttkthemes import ThemedTk
import subprocess
import json
import os
import time
import threading
import webbrowser
import sys
import numpy as np
from scipy import stats
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


#todo update_status_label einfügen
#todo wegen threading kann durch drücken prozess beliebig oft gestartet werden

# Absoluter Pfad zur JSON-Datei
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
JSON_PATH = os.path.join(PROJECT_DIR, "json_files", "simulation_data.json")
EXCEL_PATH = os.path.join(PROJECT_DIR, "analysis_results", "simulation_ergebnisse.xlsx")
JSON_INPUT = os.path.join(PROJECT_DIR, "r_code", "r_input.json")

# Simulationsresultate
all_results = []
all_train_results = []
inspector_metrics = ["inspected_trains",
                     "total_inspection_time",
                     "time_in_siding",
                     "total_pause_time"]
train_metrics = ["time",
                 "waiting_time",
                 "true_damages",
                 "true_damages_found_ai",
                 "false_positives_ai",
                 "false_negatives_ai",
                 "true_damages_found_human",
                 "false_positives_umklassifiziert",
                 "false_negatives_umklassifiziert"]

def create_slider(frame, label, from_, to, resolution, var, is_int=True):
    ttk.Label(frame, text=label).pack(anchor="w")

    # Label zur Anzeige des aktuellen Werts
    default = var.get()
    value_label = ttk.Label(frame, text=f"{default}")
    value_label.pack(anchor="center")

    def update_label(val):
        if is_int:
            value_label.config(text=f"{int(float(val))}")
        else:
            value_label.config(text=f"{float(val)}")

    slider = tk.Scale(
        frame,
        from_=from_,
        to=to,
        orient="horizontal",
        variable=var,
        resolution=resolution,
        tickinterval=0,
        showvalue=False,
        command=update_label,
        length=400
    )
    slider.pack(padx=10, pady=5)


# Funktion zum Speichern der Slider-Daten
def save_slider_data():
    try:
        with open(JSON_PATH, "w") as f:
            data = {
                "SIM_TIME": int(sim_time.get()),
                "WORKING_DAY": int(working_day.get()),
                "NUM_INSPECTORS": int(num_inspectors.get()),
                "INSP_TIME_PER_AXES": insp_time_per_axes.get(),
                "SD_INSP_TIME_PER_AXES": sd_insp_time_per_axes.get(),
                "INSP_TIME_SCREEN_PER_WAGON": insp_time_screen_per_wagon.get(),
                "SD_INSP_TIME_SCREEN_PER_WAGON": sd_insp_time_screen_per_wagon.get(),
                "INSP_CLOSER_LOOK": insp_closer_look.get(),
                "SD_INSP_CLOSER_LOOK": sd_insp_closer_look.get(),
                "TIME_PVG": time_pvg.get(),
                "SD_TIME_PVG": sd_time_pvg.get(),
                "FORMALITIES_BASELINE": formalities_baseline.get(),
                "MEAN_NUM_FORMAL_ACTS": mean_num_formal_acts.get(),
                "MEAN_TIME_FORMALS": mean_time_formals.get(),
                "SD_TIME_FORMALS": sd_time_formals.get(),
                "HUMAN_INSP_PROB": human_insp_prob.get(),
                "TRUST_AI_PROB": trust_ai_prob.get(),
                "PROB_INCON_HANDLING": prob_incon_handling.get(),
                "SHORT_PAUSE_MIN": int(short_pause_min.get()),
                "SHORT_PAUSE_MAX": int(short_pause_max.get()),
                "REGULAR_PAUSE": int(regular_pause.get()),
                "NUM_SHIFTS": int(num_shifts.get()),
                "FALSE_NEGATIVE": false_negative.get(),
                "FALSE_POSITIVE": false_positive.get(),
                "simulation_distances": {"gleise": [500, 500], "zentrale": [500, 500]}
            }
            json.dump(data, f)
        print("Slider Data erfolgreich geschrieben.")
        print(json.dumps(data, indent=4))
        status_labels["settings"].config(text="Slider Data erfolgreich gespeichert.", foreground="#0c990f")
    except Exception as e:
        print("Fehler beim Schreiben von Slider Data:", e)
        status_labels["settings"].config(text=f"Fehler: {e}", foreground="red")


# Funktion zum Starten des nächsten UI-Skripts
def start_next_ui():
    try:
        update_status_label("creator", "Creator wird gestartet...", "#0c990f")
        thread = threading.Thread(target=run_ui1_and_enable_sim_button)
        thread.start()
    except Exception as e:
        status_labels["creator"].config(text=f"Fehler: {e}", foreground="red")


def run_ui1_and_enable_sim_button():
    python_executable = sys.executable  # das ist der aktive Interpreter aus .venv
    script_path = os.path.join(os.path.dirname(__file__), "UI_1.py")
    subprocess.run([python_executable, script_path])
    root.after(0, lambda: sim_button.state(["!disabled"]))
    root.after(0, lambda: update_status_label("creator", "Werte gespeichert", "#0c990f"))


def update_status_label(key, text, color):
    if key in status_labels:
        status_labels[key].config(text=text, foreground=color)


# Öffne Simulationsfenster
def open_simulation_window():
    global settings_window
    settings_window = tk.Toplevel(root)
    settings_window.title("Einstellungen")
    settings_window.geometry("700x700")

    # Canvas und Scrollbar
    canvas = tk.Canvas(settings_window)
    scrollbar = ttk.Scrollbar(settings_window, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    status_labels["sim"].config(text="...LÄUFT...", foreground="#1911ab")

    create_slider(scrollable_frame, "Anzahl Simulationen", 0, 10000, 1, num_simulation_steps)
    start_simulation_button = ttk.Button(scrollable_frame, text="Start Simulation", command=start_simulation)
    start_simulation_button.pack(pady=20)


# Starte Simulation
def start_simulation():

    def cancel():
        cancel_simulation.set()
        if "sim" in status_labels:
            status_labels["sim"].config(text="Simulation abgebrochen", foreground="orange")
        if progress_window.winfo_exists():
            progress_window.destroy()
        if settings_window and settings_window.winfo_exists():
            settings_window.destroy()

    def run_simulation():

        try:
            num_steps = int(num_simulation_steps.get())
            progressbar["maximum"] = num_steps
            progressbar["value"] = 0

            start_time = time.time()

            for i in range(num_steps):
                if cancel_simulation.is_set():
                    return  # Beende den Thread sauber

                subprocess.run([sys.executable, "py_code/main.py", str(i)], check=True)

                # Lade JSON-Ergebnisse
                json_filename = f"results_step_{i}.json"
                with open(json_filename, "r") as f:
                    combined_results = json.load(f)

                #Wagenmeisterdaten
                step_results = combined_results["inspectors"]
                #Zug-Daten
                step_trains_results = combined_results["trains"]

                all_results.append(step_results)
                all_train_results.append(step_trains_results)

                # Lösche die Datei nach dem Einlesen
                os.remove(json_filename)

                if cancel_simulation.is_set():
                    return  # Nochmals prüfen nach dem Schritt

                # GUI-Updates nur, wenn Fenster noch existiert
                if progress_window.winfo_exists():
                    progressbar["value"] = i + 1
                    percent = int((i + 1) / num_steps * 100)
                    percent_label.config(text=f"{percent}%")
                    progress_window.update_idletasks()

            # Aggregation nach der Schleife
            inspectors_num = len(all_results[0])
            aggregated = {metric: [[] for _ in range(inspectors_num)] for metric in inspector_metrics}

            for step in all_results:
                for inspector_data in step:
                    for metric in inspector_metrics:
                        aggregated[metric][inspector_data["id"]].append(inspector_data[metric])

            # Nur einmalige Auswertung am Ende
            zeige_auswertung(aggregated, num_steps, all_train_results)

            if not cancel_simulation.is_set():
                end_time = time.time()
                duration = end_time - start_time

                if progress_window.winfo_exists():
                    progress_label.config(text=f"Fertig in {duration:.2f} Sekunden")

                if "sim" in status_labels:
                    status_labels["sim"].config(
                        text=f"Simulation durchgeführt in {duration:.2f} Sekunden",
                        foreground="#0c990f"
                    )

                # Fenster schließen nach erfolgreicher Simulation
                if progress_window.winfo_exists():
                    progress_window.destroy()
                if settings_window and settings_window.winfo_exists():
                    settings_window.destroy()

        except Exception as e:
            if "sim" in status_labels:
                status_labels["sim"].config(text=f"Fehler: {e}", foreground="red")

    # Neues Fenster für Progressbar
    progress_window = tk.Toplevel(root)
    progress_window.title("Simulationsfortschritt")
    progress_window.geometry("400x200")
    progress_label = ttk.Label(progress_window, text="Simulation läuft...")
    progress_label.pack(pady=10)
    progressbar = ttk.Progressbar(progress_window, orient="horizontal", length=300, mode="determinate")
    progressbar.pack(pady=5)
    percent_label = ttk.Label(progress_window, text="0%")
    percent_label.pack()

    cancel_simulation = threading.Event()
    cancel_button = ttk.Button(progress_window, text="Abbrechen", command=cancel)
    cancel_button.pack(pady=10)

    threading.Thread(target=run_simulation).start()


def open_choice_window():
    def on_selection(event=None):
        selected = combo.get()
        choice_window.destroy()
        run_rmarkdown(selected)

    choice_window = tk.Toplevel(root)
    choice_window.title("Auswahl der Datensätze")
    choice_window.geometry("700x700")

    label = tk.Label(choice_window, text="Wähle einen Datensatz:")
    label.pack(pady=10)

    options = ['SEE', 'BSS', 'MN', 'TK']
    combo = ttk.Combobox(choice_window, values=options, state="readonly")

    combo.set("BSS")  # Standardwert
    combo.pack(pady=10)

    button = tk.Button(choice_window, text="Auswahl bestätigen", command=on_selection)
    button.pack(pady=20)


# Funktion zum Öffnen des Einstellungsfensters
def open_settings_window():
    settings_window = tk.Toplevel(root)
    settings_window.title("Einstellungen")
    settings_window.geometry("700x700")

    # Canvas und Scrollbar
    canvas = tk.Canvas(settings_window)
    scrollbar = ttk.Scrollbar(settings_window, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Block 1
    block1 = ttk.LabelFrame(scrollable_frame, text="Block 1: Allgemeine Einstellungen")
    block1.pack(fill="x", padx=10, pady=10)

    create_slider(block1, "Simulationszeit [min]", 60, 2880, 60, sim_time)
    create_slider(block1, "Arbeitstag [min]", 60, 2880, 10, working_day)
    create_slider(block1, "Anzahl Wagenmeister", 1, 10, 1, num_inspectors)

    # Block 2
    block2 = ttk.LabelFrame(scrollable_frame, text="Block 2: Menschliche Inspektion")
    block2.pack(fill="x", padx=10, pady=10)

    create_slider(block2, "INSP_TIME_PER_AXES", 0.1, 1.0, 0.01, insp_time_per_axes, False)
    create_slider(block2, "SD_INSP_TIME_PER_AXES", 0.01, 1.0, 0.01, sd_insp_time_per_axes, False)
    create_slider(block2, "INSP_TIME_SCREEN_PER_WAGON", 0.5, 5.0, 0.1, insp_time_screen_per_wagon, False)
    create_slider(block2, "SD_INSP_TIME_SCREEN_PER_WAGON", 0.01, 1.0, 0.01, sd_insp_time_screen_per_wagon, False)
    create_slider(block2, "INSP_CLOSER_LOOK", 0.5, 5.0, 0.1, insp_closer_look, False)
    create_slider(block2, "SD_INSP_CLOSER_LOOK", 0.1, 1.0, 0.1, sd_insp_closer_look, False)
    create_slider(block2, "TIME_PVG", 1, 10, 1, time_pvg, False)
    create_slider(block2, "SD_TIME_PVG", 0.5, 5.0, 0.1, sd_time_pvg, False)
    create_slider(block2, "FORMALITIES_BASELINE", 1, 20, 1, formalities_baseline)
    create_slider(block2, "MEAN_NUM_FORMAL_ACTS", 0.0, 1.0, 0.01, mean_num_formal_acts, False)
    create_slider(block2, "MEAN_TIME_FORMALS", 1, 30, 1, mean_time_formals, False)
    create_slider(block2, "SD_TIME_FORMALS", 0.5, 10, 0.5, sd_time_formals, False)
    create_slider(block2, "HUMAN_INSP_PROB", 0.0, 1.0, 0.01, human_insp_prob, False)
    create_slider(block2, "TRUST_AI_PROB", 0.0, 1.0, 0.01, trust_ai_prob, False)
    create_slider(block2, "PROB_INCON_HANDLING", 0.0, 1.0, 0.01, prob_incon_handling, False)
    create_slider(block2, "SHORT_PAUSE Min", 1, 10, 1, short_pause_min)
    create_slider(block2, "SHORT_PAUSE Max", 10, 30, 1, short_pause_max)
    create_slider(block2, "REGULAR_PAUSE", 10, 60, 1, regular_pause)
    create_slider(block2, "NUM_SHIFTS", 1, 5, 1, num_shifts)

    # Block 3
    block3 = ttk.LabelFrame(scrollable_frame, text="Block 3: KI-Inspektion")
    block3.pack(fill="x", padx=10, pady=10)

    create_slider(block3, "FALSE_NEGATIVE", 0.0, 1, 0.01, false_negative, False)
    create_slider(block3, "FALSE_POSITIVE", 0.0, 1, 0.01, false_positive, False)

    # Speichern & Schließen Button
    def save_and_close():
        save_slider_data()
        settings_window.destroy()

    save_button = ttk.Button(scrollable_frame, text="Speichern & Schließen", command=save_and_close)
    save_button.pack(pady=20)


def run_rmarkdown(data_type="BSS"):
    r_input_data = {"type": data_type}
    os.environ["RSTUDIO_PANDOC"] = r"C:/Program Files/RStudio/resources/app/bin/quarto/bin/tools"
    output_html = r"C:/Users/rough/PycharmProjects/simpy/r_code/PVG_analysis.html"

    with open(JSON_INPUT, "w") as f:
        json.dump(r_input_data, f)

    status_labels["pvg"].config(text="Analyse läuft...", foreground="#1911ab")

    try:
        subprocess.run([
            r"C:\Program Files\R\R-4.5.1\bin\Rscript.exe",
            "--vanilla",
            "-e",
            "Sys.setenv(RSTUDIO_PANDOC=Sys.getenv('RSTUDIO_PANDOC')); "
            f"rmarkdown::render('C:/Users/rough/PycharmProjects/simpy/r_code/PVG_analysis.Rmd')"
        ], check=True, shell=True)
        status_labels["pvg"].config(text="Analyse beendet", foreground="#0c990f")
        print("Analyse erfolgreich abgeschlossen.")
        webbrowser.open(output_html)
    except subprocess.CalledProcessError as e:
        status_labels["pvg"].config(text=f"Fehler: {e}", foreground="red")
        print(f"Fehler bei der Analyse: {e}")


def create_button_with_label(parent, key, button_text, label_text, command, color="#ad1f2b", **kwargs):
    row = ttk.Frame(parent)
    row.pack(anchor="w", pady=(0, 10))

    button = ttk.Button(row, text=button_text, command=command, style="Big.TButton", **kwargs)
    button.pack(side="left")

    label = ttk.Label(row, text=label_text, foreground=color, style="Big.TLabel")

    # Label speichern, um es später ändern zu können
    status_labels[key] = label
    label.pack(side="left", padx=10)

    return button


# Mittelwerte und Konfidenzintervalle berechnen
def mean_ci(data, confidence=0.95):
    a = np.array(data)
    n = len(a)
    mean = np.mean(a)
    se = stats.sem(a)
    h = se * stats.t.ppf((1 + confidence) / 2., n - 1)
    return mean, mean - h, mean + h


def zeige_auswertung(aggregated, num_steps, all_train_results):
    global block3, block4
    results = []
    summary_results = []

    # Prüfen, ob nur ein Eintrag pro Inspektor vorhanden ist
    single_run = all(len(inspector_data) == 1 for metric in inspector_metrics for inspector_data in aggregated[metric])

    for metric in inspector_metrics:
        all_inspector_values = []

        for inspector_id, inspector_data in enumerate(aggregated[metric]):
            mean = round(sum(inspector_data) / len(inspector_data), 2)
            all_inspector_values.append(mean)

            if single_run:
                results.append({
                    "Metrik": metric,
                    "Inspektor": f"Inspektor {inspector_id}",
                    "Mittelwert": mean
                })
            else:
                mean_ci_val, lower, upper = mean_ci(inspector_data)
                results.append({
                    "Metrik": metric,
                    "Inspektor": f"Inspektor {inspector_id}",
                    "Mittelwert": round(mean_ci_val, 2),
                    "95%-KI Untergrenze": round(lower, 2),
                    "95%-KI Obergrenze": round(upper, 2)
                })

        # Zusammenfassung über alle Inspektoren
        if single_run:
            summary_results.append({
                "Metrik": metric,
                "Mittelwert über Inspektoren": round(sum(all_inspector_values) / len(all_inspector_values), 2)
            })
        else:
            mean_group, lower_group, upper_group = mean_ci(all_inspector_values)
            summary_results.append({
                "Metrik": metric,
                "Mittelwert über Inspektoren": round(mean_group, 2),
                "95%-KI Untergrenze": round(lower_group, 2),
                "95%-KI Obergrenze": round(upper_group, 2)
            })

    df_results = pd.DataFrame(results)  # results muss vorher definiert sein
    df_summary = pd.DataFrame(summary_results)  # summary_results muss vorher definiert sein

    #Eine Abbildung zwischen Bezeichnung und Variablenname
    metric_key_map = {
        "time": "stay_time",
        "waiting_time": "Wartezeit",
        "true_damages": "true_damages",
        "true_damages_found_ai": "ai_found",
        "false_positives_ai": "ai_fp",
        "false_negatives_ai": "ai_fn",
        "true_damages_found_human": "human_found",
        "false_positives_umklassifiziert": "human_fp_umklassifiziert",
        "false_negatives_umklassifiziert": "human_fn_umklassifiziert"
    }

    # Mittelwerte für Züge pro Step berechnen
    all_train_values = []
    for step_idx, step_trains in enumerate(all_train_results):
        step_metrics = {
            "step": step_idx,
            "num_trains": len(step_trains)  # Anzahl der Züge pro Step
        }
        for metric in train_metrics:
            key = metric_key_map.get(metric)
            values = [train[key] for train in step_trains if key in train]

            # Mittelwert nur für numerische Werte
            if values and all(isinstance(v, (int, float)) for v in values):
                mean_value = float(np.mean(values))
            else:
                mean_value = 0.0

            step_metrics[metric] = mean_value
        all_train_values.append(step_metrics)

    # DataFrame für Züge pro Step
    df_train_steps = pd.DataFrame(all_train_values)

    # Aggregierte Train Statistiken über alle Steps
    summary_results = []
    for metric in train_metrics:
        metric_values = [step[metric] for step in all_train_values]
        mean, ci_low, ci_high = mean_ci(metric_values)
        summary_results.append({
            "Metric": metric,
            "Mean": mean,
            "CI Lower": ci_low,
            "CI Upper": ci_high
        })

    df_train_summary = pd.DataFrame(summary_results)

    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        json_data = json.load(f)

    # Header-DataFrame zeilenweise (Index = Schlüssel, Spalte = 'Wert')
    header_rows = [("Anzahl der Simulationsschritte", num_steps)]

    for key, value in json_data.items():
        if key == "simulation_distances" and isinstance(value, dict):
            for subkey, subval in value.items():
                header_rows.append((f"{key}.{subkey}", subval))
        else:
            header_rows.append((key, value))

    header_df = pd.DataFrame(header_rows, columns=["Parameter", "Wert"])

    # Schreiben in Excel mit mehreren Blättern
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
        header_df.to_excel(writer, sheet_name="Simulation_Header")
        df_results.to_excel(writer, sheet_name="Einzelergebnisse_Inspektoren", index=False)
        df_summary.to_excel(writer, sheet_name="Zusammenfassung_Inspektoren", index=False)
        df_train_steps.to_excel(writer, sheet_name="ProSimulationsschritt_Züge", index=False)
        df_train_summary.to_excel(writer, sheet_name="Zusammenfassung_Züge", index=False)

    wb = load_workbook(EXCEL_PATH)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if cell.row == 1:
                    cell.font = Font(bold=True)

        # Spaltenbreiten automatisch anpassen
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            col_letter = column_cells[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(EXCEL_PATH)
    print(f"Excel-Datei erfolgreich gespeichert unter: {EXCEL_PATH}")

    def update_gui():
        # Block 4: Detailausgabe je Inspektor
        for widget in block4.winfo_children():
            widget.destroy()
        text_output = tk.Text(block4, wrap="none", height=20)
        text_output.pack(fill="both", expand=True)
        text_output.insert(tk.END, df_results.to_string(index=False))

        # Block 3: Zusammenfassung über alle Inspektoren
        for widget in block3.winfo_children():
            widget.destroy()
        text_summary = tk.Text(block3, wrap="none", height=10)
        text_summary.pack(fill="both", expand=True)
        text_summary.insert(tk.END, df_summary.to_string(index=False))

    root.after(0, update_gui)


def create_scrollable_block(parent, title):
    frame = ttk.LabelFrame(parent, text=title)
    frame.pack(fill="x", padx=10, pady=10)

    container = ttk.Frame(frame)
    container.pack(fill="both", expand=True)

    canvas = tk.Canvas(container)
    canvas.pack(side="left", fill="both", expand=True)

    # Scrollbars
    v_scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    v_scrollbar.pack(side="right", fill="y")
    h_scrollbar = ttk.Scrollbar(frame, orient="horizontal", command=canvas.xview)
    h_scrollbar.pack(side="bottom", fill="x")

    canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

    # Inneres Frame
    inner_frame = ttk.Frame(canvas)
    canvas_window = canvas.create_window((0, 0), window=inner_frame, anchor="nw")

    # Scrollregion aktualisieren
    def on_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfig(canvas_window, width=canvas.winfo_width())

    inner_frame.bind("<Configure>", on_configure)

    return inner_frame


# Hauptfenster
root = ThemedTk(theme="breeze")
root.title("Simpy Simulation Startbildschirm")
root.geometry("1200x800")

# Startwerte
#general params
sim_time = tk.DoubleVar(value=1440)
working_day = tk.DoubleVar(value=1440)
num_inspectors = tk.DoubleVar(value=3)
#Wagenmeister params
insp_time_per_axes = tk.DoubleVar(value=0.35)
sd_insp_time_per_axes = tk.DoubleVar(value=0.1)
insp_time_screen_per_wagon = tk.DoubleVar(value=1)
sd_insp_time_screen_per_wagon = tk.DoubleVar(value=0.1)
insp_closer_look = tk.DoubleVar(value=2)
sd_insp_closer_look = tk.DoubleVar(value=0.3)
time_pvg = tk.DoubleVar(value=5)
sd_time_pvg = tk.DoubleVar(value=1)
formalities_baseline = tk.DoubleVar(value=10)
mean_num_formal_acts = tk.DoubleVar(value=0.2)
mean_time_formals = tk.DoubleVar(value=20)
sd_time_formals = tk.DoubleVar(value=3)
human_insp_prob = tk.DoubleVar(value=0.99)
trust_ai_prob = tk.DoubleVar(value=0.5)
prob_incon_handling = tk.DoubleVar(value=0.5)
short_pause_min = tk.DoubleVar(value=3)
short_pause_max = tk.DoubleVar(value=20)
regular_pause = tk.DoubleVar(value=30)
num_shifts = tk.DoubleVar(value=3)
#AI params
false_negative = tk.DoubleVar(value=0.02)
false_positive = tk.DoubleVar(value=0.05)
num_simulation_steps = tk.DoubleVar(value=100)

# Canvas und Scrollbar
container = ttk.Frame(root)
container.pack(fill="both", expand=True)

canvas = tk.Canvas(container)
scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
canvas.configure(yscrollcommand=scrollbar.set)

scrollbar.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)

# Scrollbares Frame im Canvas
scrollable_frame = ttk.Frame(canvas)
scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(
        scrollregion=canvas.bbox("all"),
        width=scrollable_frame.winfo_reqwidth()
    )
)

canvas_frame = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

# Dynamische Breitenanpassung
def resize_canvas(event):
    canvas.itemconfig(canvas_frame, width=event.width)

canvas.bind("<Configure>", resize_canvas)

# Style mit größerer Schrift definieren
style = ttk.Style()
style.configure("Big.TButton", font=("Arial", 12))
style.configure("Big.TLabel", font=("Arial", 12))

# Dictionary zum Speichern der Labels
status_labels = {}

# Block 1
block1 = ttk.LabelFrame(scrollable_frame, text="PVG-Analyse")
block1.pack(fill="x", padx=10, pady=10)
create_button_with_label(block1, "pvg", "Generiere HTML", "keine Analyse", open_choice_window)

# Block 2
block2 = ttk.LabelFrame(scrollable_frame, text="Einstellungen")
block2.pack(fill="x", padx=10, pady=10)
create_button_with_label(block2, "settings", "Parameter", "Standardwerte voreingestellt", open_settings_window, color="#0c990f")
create_button_with_label(block2, "creator", "Bahnhof-Creator", "keine Werte gespeichert", start_next_ui)
sim_button = create_button_with_label(block2, "sim", "Simulation", "keine Simulation", open_simulation_window)

# Weitere Blöcke
block3 = create_scrollable_block(scrollable_frame, "Zusammenfassung")
block4 = create_scrollable_block(scrollable_frame, "Zusammenfassung pro Inspektor")

root.mainloop()
